import os
import json
import polars as pl
from openpyxl import Workbook
from xml.sax.saxutils import escape
from .exceptions import WriterError
from typing import List, Dict, Any, Optional

EXCEL_MAX_ROWS = 1_048_000  # safe threshold

def ensure_parent(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)

def write_csv(df: pl.DataFrame, path: str):
    if not isinstance(df, pl.DataFrame):
        raise WriterError("Input must be a Polars DataFrame")
    if not path or not isinstance(path, str):
        raise WriterError("Path must be a non-empty string")
    
    ensure_parent(path)
    try:
        df.write_csv(path)
    except Exception as e:
        raise WriterError(f"Failed to write CSV: {e}") from e

def write_ndjson(df: pl.DataFrame, path: str):
    ensure_parent(path)
    try:
        df.write_ndjson(path)
    except Exception as e:
        raise WriterError(f"Failed to write line-delimited JSON: {e}") from e

def write_json(df: pl.DataFrame, path: str):
    """Write DataFrame as traditional JSON array format"""
    ensure_parent(path)
    try:
        # Convert to list of dictionaries
        data = df.to_dicts()
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    except Exception as e:
        raise WriterError(f"Failed to write JSON array: {e}") from e

def write_xlsx(df: pl.DataFrame, path: str):
    if not isinstance(df, pl.DataFrame):
        raise WriterError("Input must be a Polars DataFrame")
    if not path or not isinstance(path, str):
        raise WriterError("Path must be a non-empty string")
    
    ensure_parent(path)
    try:
        total = df.height
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        start = 0
        sheet_idx = 1
        while start < total:
            end = min(start + EXCEL_MAX_ROWS, total)
            chunk = df.slice(start, end - start)
            
            # Create new sheet
            ws = wb.create_sheet(f"Sheet{sheet_idx}")
            
            # Write headers
            for col_idx, col_name in enumerate(chunk.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Write data
            for row_idx, row in enumerate(chunk.iter_rows(named=True), 2):
                for col_idx, col_name in enumerate(chunk.columns, 1):
                    value = row[col_name]
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            start = end
            sheet_idx += 1
        
        wb.save(path)
    except Exception as e:
        raise WriterError(f"Failed to write XLSX: {e}") from e

def write_xml(df: pl.DataFrame, path: str, root_tag: str = "records", row_tag: str = "record"):
    ensure_parent(path)
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(f"<{root_tag}>\n")
            cols = df.columns
            for row in df.iter_rows(named=True):
                f.write(f"  <{row_tag}>")
                for c in cols:
                    v = "" if row[c] is None else str(row[c])
                    f.write(f"<{c}>{escape(v)}</{c}>")
                f.write(f"</{row_tag}>\n")
            f.write(f"</{root_tag}>\n")
    except Exception as e:
        raise WriterError(f"Failed to write XML: {e}") from e

def write_positional(df: pl.DataFrame, path: str, mappings: List[Dict[str, Any]], logger: Optional[Any] = None):
    ensure_parent(path)
    targets = [m["target"] for m in mappings if "target" in m]
    lengths = [m.get("length") for m in mappings if "target" in m]
    try:
        with open(path, "w", encoding="utf-8") as f:
            for ridx, row in enumerate(df.iter_rows(named=True)):
                pieces = []
                for t, L in zip(targets, lengths):
                    val = row.get(t, "")
                    s = "" if val is None else str(val)
                    width = int(L) if L is not None else len(s)
                    if width < len(s):
                        if logger:
                            logger.warning(f"Truncating column '{t}' at row {ridx}: '{s}' -> width {width}")
                        s = s[:width]
                    # right-align numeric
                    try:
                        float(s)
                        aligned = s.rjust(width)
                    except:
                        aligned = s.ljust(width)
                    pieces.append(aligned)
                f.write("".join(pieces) + "\n")
    except Exception as e:
        raise WriterError(f"Failed to write positional: {e}") from e

def write_output(df: pl.DataFrame, base_path: str, fmt: str, mappings: List[Dict[str, Any]], xml_cfg: Optional[Dict[str, Any]] = None, logger: Optional[Any] = None) -> str:
    fmt = fmt.lower()
    if fmt == "csv":
        out_path = f"{base_path}.csv"
        write_csv(df, out_path)
        return out_path
    if fmt == "json":
        out_path = f"{base_path}.jsonl"
        write_ndjson(df, out_path)
        return out_path
    if fmt == "json_array":
        out_path = f"{base_path}.json"
        write_json(df, out_path)
        return out_path
    if fmt == "xlsx":
        out_path = f"{base_path}.xlsx"
        write_xlsx(df, out_path)
        return out_path
    if fmt == "xml":
        out_path = f"{base_path}.xml"
        root_tag = (xml_cfg or {}).get("root_tag", "records")
        row_tag = (xml_cfg or {}).get("row_tag", "record")
        write_xml(df, out_path, root_tag, row_tag)
        return out_path
    if fmt == "positional":
        out_path = f"{base_path}.txt"
        write_positional(df, out_path, mappings, logger=logger)
        return out_path
    raise WriterError(f"Unsupported output_format: {fmt}")

